import pandas as pd
import fitz
import cv2
from io import BytesIO
from PIL import Image
import base64
from openai import OpenAI
from pydantic import BaseModel
import json
import numpy as np
import traceback
from typing import Optional
import time
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Border, Side, Font

from blank_functions.form_class import Form

promt = """"You are tasked with extracting information from an image of a completed exam answer sheet and converting it into a structured JSON format. The answer sheet has the following structure:

Top Section:

Document Title: Labeled as "БЛАНК ОТВЕТОВ."
Subject Name: Handwritten in a grid labeled "МАТЕМАТИКА."
Participant Code: A field containing four cells where numbers are written.
Version Number: A field containing four cells where numbers are written.
Instructions: Indicates that answers can be integers or finite decimal fractions with 1-2 decimal places. It specifies that units of measurement and periods should not be included after the answer.
Example: The example answer "-0,9" is provided.
Middle Section:

Filling Instructions: Explains that filling should be done using a black gel or capillary pen with digits of a specific standard. Acceptable symbols are listed as: ", -1234567890."
Example of Valid Symbols: Includes symbols such as ", - 1 2 3 4 5 6 7 8 9 0 . "
Main Section: Divided into two large columns:

Left Column: "Ответы к заданиям"

Numbered from 1 to 10.
Each question is accompanied by a grid of cells to record the answer.
Examples of filled answers:
№1: "275"
№2: "38"
№3: "14"
№4: "3478"
№5: "102"
№6: "3"
№7: "0,5"
№8: "18"
Right Column: "Замена ошибочных ответов на задания"

Numbered similarly to the left column (1–10).
Intended for recording corrected answers for mistakes.
Examples of corrections:
№1: Corrected answer "-275."
№7: Corrected answer "-3."
Your task:

Extract all the information from the described answer sheet and format it in a JSON structure.
Ensure the JSON includes all relevant details such as subject name, participant code, version number, and both the answers and corrections for questions.
Example JSON Structure:
{
  "subject_name": "МАТЕМАТИКА",
  "participant_code": "1234",
  "version_number": "5678",
  "answer_1": "275",
  "answer_2": "38",
  "answer_3": "14",
  "answer_4": "3478",
  "answer_5": "102",
  "answer_6": "3",
  "answer_7": "0,5",
  "answer_8": "18",
  "answer_9": "",
  "answer_10": "",
  "correction_1": "-275",
  "correction_2": "",
  "correction_3": "",
  "correction_4": "",
  "correction_5": "",
  "correction_6": "",
  "correction_7": "-3",
  "correction_8": "",
  "correction_9": "",
  "correction_10": ""
}

Be careful! Pay special attention to the minus sign and the comma separately. Don't miss them!
"""

#
def encode_image(image):
    image = Image.fromarray(image)

    # Save the image to a BytesIO buffer in PNG format
    buffer = BytesIO()
    image.save(buffer, format="PNG")
    buffer.seek(0)

    # Encode the PNG image to base64
    test_pic_bytes = buffer.getvalue()
    base64_encoded = base64.b64encode(test_pic_bytes).decode('utf-8')

    return base64_encoded

def extract_text_from_image(api_key, image, prompt):
    client = OpenAI(api_key=api_key)

    # Getting the base64 string
    base64_image = encode_image(image)

    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": prompt
                    },
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/jpeg;base64,{base64_image}",
                            "detail": "high"
                        },
                    },
                ],
            }
        ],
        response_format={
                'type': 'json_schema',
                'json_schema':
                    {
                        "name": "whocares",
                        "schema": ResearchPaperExtraction.model_json_schema()
                    }},
    )

    json_response = json.loads(response.choices[0].message.content)
    return json_response

def get_pic_from_pdf(pdf_stream, index, zoom=6.0):
    """
    Получает изображение страницы из PDF с высоким качеством
    :param pdf_stream: поток PDF файла
    :param index: индекс страницы
    :param zoom: коэффициент увеличения (по умолчанию 4.0 для ~600 DPI)
    :return: изображение в формате NumPy (серое)
    """
    # Открываем PDF из байтового потока
    start_time = time.time()
    pdf_document = fitz.open(stream=pdf_stream, filetype="pdf")
    time_1 = time.time()
    page = pdf_document.load_page(index)  # Загружаем страницу
    time_2 = time.time()
    # Матрица для увеличения качества (увеличение DPI)
    mat = fitz.Matrix(zoom, zoom)
    time_3 = time.time()
    # Преобразуем страницу в изображение
    pix = page.get_pixmap(matrix=mat, colorspace=fitz.csRGB)  # Улучшенное цветовое пространство
    time_4 = time.time()
    # Конвертируем Pixmap в Pillow Image
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    time_5 = time.time()
    # Опционально: сохраняем в PNG для максимального качества перед конвертацией
    # img.save("original_image.png", format="PNG")
    time_6 = time.time()
    # Конвертируем в оттенки серого с максимальным качеством
    img_gray = cv2.cvtColor(np.array(img), cv2.COLOR_RGB2GRAY)
    time_7 = time.time()

    return img_gray

def get_correct_answers(correct_answers_path):
    correct_answers = pd.read_excel(correct_answers_path)
    for col in correct_answers.columns:
        if col != 'Вариант':
            correct_answers[col] = correct_answers[col].astype(str)
            correct_answers[col] = correct_answers[col].apply(lambda x: x.replace(',', '.'))
            correct_answers[col] = correct_answers[col].astype(float)
            correct_answers[col] = correct_answers[col].astype(str)
    correct_answers['Вариант'] = correct_answers['Вариант'].astype(str)
    correct_answers.rename(columns={1: "Правильный ответ 1", 2: "Правильный ответ 2", 3: "Правильный ответ 3", 4: "Правильный ответ 4", 5: "Правильный ответ 5", 6: "Правильный ответ 6", 7: "Правильный ответ 7", 8: "Правильный ответ 8", 9: "Правильный ответ 9", 10: "Правильный ответ 10"}, inplace=True)
    return correct_answers

def postprocess_raw_output(df_global_fin, correct_answers):
    df_global_fin['Предмет'] = df_global_fin['Предмет'].str.upper()
    df_global_fin['Вариант'] = df_global_fin['Вариант'].astype(int)
    df_global_fin['Вариант'] = df_global_fin['Вариант'].astype(str)
    for i in range(1, 11):
        df_global_fin[f'Задание {i}'] = df_global_fin[f'Задание {i}'].astype(float)
        df_global_fin[f'Замена {i}'] = df_global_fin[f'Замена {i}'].astype(float)
    for col in df_global_fin.columns:
        df_global_fin[col] = df_global_fin[col].astype(str)

    total_df = pd.merge(df_global_fin, correct_answers, on="Вариант", how="left")

    return total_df

class ResearchPaperExtraction(BaseModel):
    subject_name: str
    participant_code: int
    version_number: int
    answer_1: Optional[float]
    answer_2: Optional[float]
    answer_3: Optional[float]
    answer_4: Optional[float]
    answer_5: Optional[float]
    answer_6: Optional[float]
    answer_7: Optional[float]
    answer_8: Optional[float]
    answer_9: Optional[float]
    answer_10: Optional[float]
    correction_1: Optional[float]
    correction_2: Optional[float]
    correction_3: Optional[float]
    correction_4: Optional[float]
    correction_5: Optional[float]
    correction_6: Optional[float]
    correction_7: Optional[float]
    correction_8: Optional[float]
    correction_9: Optional[float]
    correction_10: Optional[float]

# def transform_json_to_dataframe(parsed_json, answer_minus_list, correction_minus_list, row_image):
#   df = pd.DataFrame([parsed_json])
#   order = ['subject_name', 'participant_code', 'version_number', 'answer_1', 'answer_2', 'answer_3', 'answer_4', 'answer_5', 'answer_6', 'answer_7', 'answer_8', 'answer_9', 'answer_10', 'correction_1', 'correction_2', 'correction_3', 'correction_4', 'correction_5', 'correction_6', 'correction_7', 'correction_8', 'correction_9', 'correction_10']
#   df = df[order]
#   rename_dict = {
#     'subject_name': 'Предмет',
#     'participant_code': 'Код участника',
#     'version_number': 'Вариант',
#     'answer_1': 'Задание 1',
#     'answer_2': 'Задание 2',
#     'answer_3': 'Задание 3',
#     'answer_4': 'Задание 4',
#     'answer_5': 'Задание 5',
#     'answer_6': 'Задание 6',
#     'answer_7': 'Задание 7',
#     'answer_8': 'Задание 8',
#     'answer_9': 'Задание 9',
#     'answer_10': 'Задание 10',
#     'correction_1': 'Замена 1',
#     'correction_2': 'Замена 2',
#     'correction_3': 'Замена 3',
#     'correction_4': 'Замена 4',
#     'correction_5': 'Замена 5',
#     'correction_6': 'Замена 6',
#     'correction_7': 'Замена 7',
#     'correction_8': 'Замена 8',
#     'correction_9': 'Замена 9',
#     'correction_10': 'Замена 10',
#     'row_image_1': 'Картинка ответа 1',
#     'row_image_2': 'Картинка ответа 2',
#     'row_image_3': 'Картинка ответа 3',
#     'row_image_4': 'Картинка ответа 4',
#     'row_image_5': 'Картинка ответа 5',
#     'row_image_6': 'Картинка ответа 6',
#     'row_image_7': 'Картинка ответа 7',
#     'row_image_8': 'Картинка ответа 8',
#     'row_image_9': 'Картинка ответа 9',
#     'row_image_10': 'Картинка ответа 10'
#   }
  
#   for i in range(1, 11):
#       df[f'answer_{i}'] = df[f'answer_{i}'].apply(lambda x: abs(float(x)) * answer_minus_list[i - 1] if x != None else None)
#       df[f'correction_{i}'] = df[f'correction_{i}'].apply(lambda x: abs(float(x)) * correction_minus_list[i - 1] if x != None else None)
#       df[f'row_image_{i}'] = [row_image]
#   df.rename(columns=rename_dict, inplace=True)
#   return df

def transform_json_to_dataframe(parsed_json, answer_minus_list, correction_minus_list):
  df = pd.DataFrame([parsed_json])
  order = ['subject_name', 'participant_code', 'version_number', 'answer_1', 'answer_2', 'answer_3', 'answer_4', 'answer_5', 'answer_6', 'answer_7', 'answer_8', 'answer_9', 'answer_10', 'correction_1', 'correction_2', 'correction_3', 'correction_4', 'correction_5', 'correction_6', 'correction_7', 'correction_8', 'correction_9', 'correction_10']
  df = df[order]
  rename_dict = {
    'subject_name': 'Предмет',
    'participant_code': 'Код участника',
    'version_number': 'Вариант',
    'answer_1': 'Задание 1',
    'answer_2': 'Задание 2',
    'answer_3': 'Задание 3',
    'answer_4': 'Задание 4',
    'answer_5': 'Задание 5',
    'answer_6': 'Задание 6',
    'answer_7': 'Задание 7',
    'answer_8': 'Задание 8',
    'answer_9': 'Задание 9',
    'answer_10': 'Задание 10',
    'correction_1': 'Замена 1',
    'correction_2': 'Замена 2',
    'correction_3': 'Замена 3',
    'correction_4': 'Замена 4',
    'correction_5': 'Замена 5',
    'correction_6': 'Замена 6',
    'correction_7': 'Замена 7',
    'correction_8': 'Замена 8',
    'correction_9': 'Замена 9',
    'correction_10': 'Замена 10',
    'row_image_1': 'Картинка ответа 1',
    'row_image_2': 'Картинка ответа 2',
    'row_image_3': 'Картинка ответа 3',
    'row_image_4': 'Картинка ответа 4',
    'row_image_5': 'Картинка ответа 5',
    'row_image_6': 'Картинка ответа 6',
    'row_image_7': 'Картинка ответа 7',
    'row_image_8': 'Картинка ответа 8',
    'row_image_9': 'Картинка ответа 9',
    'row_image_10': 'Картинка ответа 10'
  }
  
  for i in range(1, 11):
      df[f'answer_{i}'] = df[f'answer_{i}'].apply(lambda x: float(x) if x != None else None)
      df[f'correction_{i}'] = df[f'correction_{i}'].apply(lambda x: float(x) if x != None else None)
    #   df[f'row_image_{i}'] = [row_image]
  df.rename(columns=rename_dict, inplace=True)
  return df

def check_answers(total_df):

    scores_dict = {
    1: 0.5,
    2: 1,
    3: 1,
    4: 1,
    5: 1,
    6: 1,
    7: 1,
    8: 1,
    9: 1,
    10: 1.5
    }
    # for i in range(1, 11):
    #     right_answer_v2 = total_df[f'Правильный ответ {i}'].apply(lambda x: x.replace('.', '1') + '.0')
    #     print(right_answer_v2)
    #     print(total_df[f'Задание {i}'])
    #     print(total_df[f'Замена {i}'])
    #     if right_answer_v2.values[0] == total_df[f'Задание {i}'].values[0]:
    #         total_df[f'Задание {i}'] = total_df[f'Правильный ответ {i}']
    #     if right_answer_v2.values[0] == total_df[f'Замена {i}'].values[0]:
    #         total_df[f'Замена {i}'] = total_df[f'Правильный ответ {i}']

    for i in range(1, 11):
        total_df[f'Начисленные баллы {i}'] = ((total_df[f'Правильный ответ {i}'] == total_df[f'Задание {i}']) | (total_df[f'Правильный ответ {i}'] == total_df[f'Замена {i}'])).replace(True, 'Верно').replace(False, 'Неверно').apply(lambda x: scores_dict[i] if x == 'Верно' else 0)

    total_df['Начисленные баллы сумма'] = total_df[[f'Начисленные баллы {i}' for i in range(1, 11)]].sum(axis=1)

    return total_df

def final_styling(total_df):
    reorder_cols_list = ['Предмет', 'Код участника', 'Вариант']
    for i in range(1, 11):  
        reorder_cols_list.append(f'Задание {i}')
        reorder_cols_list.append(f'Замена {i}')
        reorder_cols_list.append(f'Картинка ответа {i}')

    for i in range(1, 11):
        reorder_cols_list.append(f'Начисленные баллы {i}')
    reorder_cols_list.append('Начисленные баллы сумма')
    total_df = total_df[reorder_cols_list]

    for i in range(1, 11):
        total_df.loc[:, f'Задание {i}'] = total_df[f'Замена {i}'].where(total_df[f'Замена {i}'] != 'nan', total_df[f'Задание {i}'])
        total_df = total_df.drop(columns=[f'Замена {i}'])

    for col in total_df.columns:
        total_df[col] = total_df[col].replace('nan', '')

    return total_df

def save_to_excel(df_global_styled, file_name="Formatted_Data.xlsx"):
    columns = ["Предмет", "Код участника", "Вариант", "Задание 1", "Картинка ответа 1", "Задание 2", "Картинка ответа 2", "Задание 3", "Картинка ответа 3", "Задание 4", "Картинка ответа 4", "Задание 5", "Картинка ответа 5", "Задание 6", "Картинка ответа 6", "Задание 7", "Картинка ответа 7", "Задание 8", "Картинка ответа 8", "Задание 9", "Картинка ответа 9", "Задание 10", "Картинка ответа 10", "Начисленные баллы 1", "Начисленные баллы 2", "Начисленные баллы 3", "Начисленные баллы 4", "Начисленные баллы 5", "Начисленные баллы 6", "Начисленные баллы 7", "Начисленные баллы 8", "Начисленные баллы 9", "Начисленные баллы 10"]

    # Save to Excel with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "Общая таблица"

    # Write the header
    header = ["Предмет", "Код участника", "Вариант", "Ответы", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, "Баллы", None, None, None, None, None, None, None, None, None, "Всего"]
    ws.append(header)
    ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=23)
    ws.merge_cells(start_row=1, start_column=24, end_row=1, end_column=33)
    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=3, end_column=2)
    ws.merge_cells(start_row=1, start_column=3, end_row=3, end_column=3)
    ws.merge_cells(start_row=1, start_column=34, end_row=2, end_column=34)

    # Write the subheader
    ws.append(columns)

    # Write the data
    for row in dataframe_to_rows(df_global_styled, index=False, header=False):
        ws.append(row)

    # Align headers
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Create a border for the entire dataframe
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Make the first three rows bold
    for row in ws.iter_rows(min_row=1, max_row=3):
        for cell in row:
            cell.font = Font(bold=True)

    for col in range(4, 24):
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)

    # Save to a BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output